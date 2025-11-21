"""DocFlow CLI entrypoint."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional, List

import typer
import click
from openpyxl import Workbook
from openpyxl.styles import Font

from docflow.core.extraction.engine import ExtractionResult, MultiResult
from docflow.core.errors import DocumentError, ExtractionError, ProviderError
from docflow.core.utils.io import load_structured
from docflow.sdk.client import DocflowClient
from docflow.sdk.config import DEFAULT_CONFIG_PATH, SdkConfig, load_config, merge_cli_overrides
from docflow.sdk.errors import ConfigError, RemoteServiceError
from docflow.sdk import profiles

app = typer.Typer(add_completion=False, help="DocFlow CLI")


class Context:
    def __init__(self) -> None:
        self.config = load_config()
        self.output_format = self.config.default_output_format
        self.output_path: Optional[Path] = None
        self.multi = self.config.mode  # placeholder, overwritten per command
        self.verbose = False


# --- utility helpers ---

def _result_to_obj(result: Any) -> Any:
    if isinstance(result, MultiResult):
        return result.to_dict()
    if isinstance(result, ExtractionResult):
        return result.to_dict()
    if isinstance(result, list):
        return [_result_to_obj(r) for r in result]
    return result


def _ensure_directory(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _write_json(path: Path, payload: Any) -> None:
    _ensure_directory(path)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _export_excel_single(data: dict, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GlobalFields"
    bold = Font(bold=True)

    global_fields = {k: v for k, v in data.items() if not isinstance(v, list)}
    row = 1
    for name, value in global_fields.items():
        ws.cell(row=row, column=1, value=name).font = bold
        ws.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value)
        row += 1

    for name, records in data.items():
        if not isinstance(records, list):
            continue
        sheet = wb.create_sheet(f"RS_{name}")
        headers = set()
        for rec in records:
            if isinstance(rec, dict):
                headers.update(rec.keys())
        headers_sorted = sorted(headers)
        for col, header in enumerate(headers_sorted, start=1):
            sheet.cell(row=1, column=col, value=header).font = bold
        for r_idx, rec in enumerate(records, start=2):
            if not isinstance(rec, dict):
                continue
            for c_idx, header in enumerate(headers_sorted, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=rec.get(header))
    _ensure_directory(path)
    wb.save(path)


def _handle_excel(result: Any, output_path: Path | None) -> None:
    if isinstance(result, MultiResult):
        per_file = result.per_file
        aggregate = result.aggregate
        for idx, item in enumerate(per_file, start=1):
            name = item.meta.get("docs", [f"doc{idx}"])[0] if isinstance(item.meta, dict) else f"doc{idx}"
            target = output_path
            if target and target.suffix.lower() == ".xlsx" and len(per_file) > 1:
                target = target.with_name(f"{target.stem}_{idx}{target.suffix}")
            if target is None or target.suffix.lower() != ".xlsx":
                target = Path.cwd() / f"docflow_{name}_{idx}.xlsx"
            _export_excel_single(item.data, target)
        if aggregate:
            target = output_path or Path.cwd() / "docflow_aggregate.xlsx"
            _export_excel_single(aggregate.data, target)
        return

    if isinstance(result, list):
        for idx, item in enumerate(result, start=1):
            if isinstance(item, ExtractionResult):
                name = item.meta.get("docs", [f"doc{idx}"])[0] if isinstance(item.meta, dict) else f"doc{idx}"
                target = output_path
                if target and target.suffix.lower() == ".xlsx" and len(result) > 1:
                    target = target.with_name(f"{target.stem}_{idx}{target.suffix}")
                if target is None or target.suffix.lower() != ".xlsx":
                    target = Path.cwd() / f"docflow_{name}_{idx}.xlsx"
                _export_excel_single(item.data, target)
        return

    if isinstance(result, ExtractionResult):
        target = output_path or Path.cwd() / "docflow_output.xlsx"
        _export_excel_single(result.data, target)
        return

    raise typer.Exit(code=1)


def _print_output(result: Any, output_format: str, output_path: Path | None) -> None:
    obj = _result_to_obj(result)
    if output_format == "print":
        typer.echo(json.dumps(obj, indent=2, ensure_ascii=False))
    elif output_format == "json":
        if output_path:
            _write_json(output_path, obj)
        else:
            typer.echo(json.dumps(obj, indent=2, ensure_ascii=False))
    elif output_format == "excel":
        _handle_excel(result, output_path)
    else:
        typer.echo(f"Unsupported output format: {output_format}")
        raise typer.Exit(code=1)


def _handle_exc(err: Exception) -> None:
    """Print a concise error and exit non-zero."""
    typer.echo(f"Error: {err}", err=True)
    raise typer.Exit(code=1)


def _make_client(ctx: Context, mode: str | None, base_url: str | None) -> DocflowClient:
    cfg = merge_cli_overrides(ctx.config, mode=mode, endpoint=base_url)
    return DocflowClient(mode=cfg.mode, endpoint_url=cfg.endpoint_url, config=cfg)


# --- CLI commands ---


@app.callback()
def main(
    ctx: typer.Context,
    verbose: bool = typer.Option(False, "--verbose", help="Verbose output"),
) -> None:
    if ctx.obj is None:
        ctx.obj = Context()
    ctx.obj.verbose = verbose


@app.command()
def init(
    ctx: typer.Context,
    base_url: str = typer.Option("", "--base-url", help="Default service endpoint"),
    default_output_format: str = typer.Option("json", "--default-output-format", help="Default output format"),
    default_output_dir: Path = typer.Option(Path("./outputs"), "--default-output-dir", help="Default output directory"),
) -> None:
    context: Context = ctx.obj
    cfg_dir = DEFAULT_CONFIG_PATH.parent
    cfg_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "docflow": {
            "mode": context.config.mode,
            "endpoint": base_url or context.config.endpoint_url,
            "default_output_format": default_output_format,
            "default_output_dir": str(default_output_dir),
        }
    }
    DEFAULT_CONFIG_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    typer.echo(f"Wrote config to {DEFAULT_CONFIG_PATH}")


@app.command()
def extract(
    ctx: typer.Context,
    schema: Optional[Path] = typer.Option(None, "--schema", help="Path to schema file"),
    all_fields: bool = typer.Option(False, "--all", help="Schema-less extraction"),
    multi: str = typer.Option("per_file", "--multi", help="per_file|aggregate|both"),
    base_url: str = typer.Option("", "--base-url", help="Remote service base URL"),
    mode: Optional[str] = typer.Option(None, "--mode", help="local or remote"),
    output_format: Optional[str] = typer.Option(None, "--output-format", help="print|json|excel"),
    output_path: Optional[Path] = typer.Option(None, "--output-path", help="Write output to file"),
    files: List[Path] = typer.Argument(..., exists=False, readable=False, help="Document files"),
) -> None:
    context: Context = ctx.obj
    if not all_fields and schema is None:
        typer.echo("--schema is required unless --all is specified", err=True)
        raise typer.Exit(code=1)
    if all_fields and schema is not None:
        typer.echo("--schema and --all are mutually exclusive", err=True)
        raise typer.Exit(code=1)

    cfg_output_format = output_format or context.config.default_output_format
    context.output_path = output_path

    client = _make_client(context, mode=mode, base_url=base_url or None)

    try:
        if all_fields:
            result = client.extract_all([str(p) for p in files], multi_mode=multi)
        else:
            schema_dict = load_structured(schema)
            result = client.extract(schema_dict, [str(p) for p in files], multi_mode=multi)
    except (ConfigError, RemoteServiceError, DocumentError, ProviderError, ExtractionError, FileNotFoundError) as exc:
        _handle_exc(exc)

    _print_output(result, cfg_output_format, output_path)


@app.command()
def describe(
    ctx: typer.Context,
    multi: str = typer.Option("per_file", "--multi", help="per_file|aggregate|both"),
    base_url: str = typer.Option("", "--base-url", help="Remote service base URL"),
    mode: Optional[str] = typer.Option(None, "--mode", help="local or remote"),
    output_format: Optional[str] = typer.Option(None, "--output-format", help="print|json|excel"),
    output_path: Optional[Path] = typer.Option(None, "--output-path", help="Write output to file"),
    files: List[Path] = typer.Argument(..., help="Document files"),
) -> None:
    context: Context = ctx.obj
    cfg_output_format = output_format or context.config.default_output_format
    client = _make_client(context, mode=mode, base_url=base_url or None)
    try:
        result = client.describe([str(p) for p in files], multi_mode=multi)
    except (ConfigError, RemoteServiceError, DocumentError, ProviderError, ExtractionError, FileNotFoundError) as exc:
        _handle_exc(exc)
    _print_output(result, cfg_output_format, output_path)


@app.command()
def run(
    ctx: typer.Context,
    profile_name: str = typer.Argument(..., help="Profile name"),
    multi: str = typer.Option("per_file", "--multi", help="per_file|aggregate|both"),
    base_url: str = typer.Option("", "--base-url", help="Remote service base URL"),
    mode: Optional[str] = typer.Option(None, "--mode", help="local or remote"),
    output_format: Optional[str] = typer.Option(None, "--output-format", help="print|json|excel"),
    output_path: Optional[Path] = typer.Option(None, "--output-path", help="Write output to file"),
    files: List[Path] = typer.Argument(..., help="Document files"),
) -> None:
    context: Context = ctx.obj
    cfg_output_format = output_format or context.config.default_output_format
    client = _make_client(context, mode=mode, base_url=base_url or None)
    try:
        result = client.run_profile(profile_name, [str(p) for p in files], multi_mode=multi)
    except (ConfigError, RemoteServiceError, DocumentError, ProviderError, ExtractionError, FileNotFoundError) as exc:
        _handle_exc(exc)
    _print_output(result, cfg_output_format, output_path)


profiles_app = typer.Typer(help="Profile utilities")


@profiles_app.command("list")
def profiles_list(ctx: typer.Context) -> None:
    context: Context = ctx.obj
    names = profiles.list_profiles(context.config)
    for name in names:
        typer.echo(name)


@profiles_app.command("show")
def profiles_show(ctx: typer.Context, profile_name: str = typer.Argument(...)) -> None:
    context: Context = ctx.obj
    profile = profiles.load_profile(profile_name, context.config)
    payload = {
        "name": profile.name,
        "mode": profile.mode,
        "multi": profile.multi_mode_default,
        "description": profile.description,
    }
    typer.echo(json.dumps(payload, indent=2))


app.add_typer(profiles_app, name="profiles")


if __name__ == "__main__":  # pragma: no cover
    app()
