"""Excel export utilities for hierarchical JSON data."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

INVALID_SHEET_CHARS = set("[]:*?/\\'")
MAX_SHEET_NAME_LEN = 31


def export_json_to_excel(data: Any, path: Path) -> None:
    exporter = _JsonToExcelExporter()
    exporter.write_workbook(data)
    path.parent.mkdir(parents=True, exist_ok=True)
    exporter.save(path)


class _JsonToExcelExporter:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.bold = Font(bold=True)
        self.sheet_names: set[str] = set()

    def save(self, path: Path) -> None:
        self.wb.save(path)

    def write_workbook(self, data: Any) -> None:
        if isinstance(data, dict):
            ws = self._init_root_sheet("OBJ", "root")
            self._write_object_sheet(ws, data, "root")
        elif isinstance(data, list):
            ws = self._init_root_sheet("ARR", "root")
            if data:
                self._write_array_sheet(ws, data, "root")
        else:
            ws = self._init_root_sheet("VAL", "root")
            ws.cell(row=1, column=1, value="Value").font = self.bold
            ws.cell(row=2, column=1, value=self._format_scalar(data))

    def _init_root_sheet(self, prefix: str, label: str):
        ws = self.wb.active
        ws.title = self._alloc_sheet_name(prefix, label)
        return ws

    def _create_sheet(self, prefix: str, label: str):
        name = self._alloc_sheet_name(prefix, label)
        return self.wb.create_sheet(title=name)

    def _alloc_sheet_name(self, prefix: str, label: str) -> str:
        base_label = self._slug(label)
        base = f"{prefix}_{base_label}" if base_label else prefix
        base = base[:MAX_SHEET_NAME_LEN]
        candidate = base
        suffix = 1
        while candidate in self.sheet_names:
            suffix_str = f"_{suffix}"
            trim_len = MAX_SHEET_NAME_LEN - len(suffix_str)
            candidate = f"{base[:trim_len]}{suffix_str}"
            suffix += 1
        self.sheet_names.add(candidate)
        return candidate

    def _slug(self, label: str) -> str:
        text = str(label)
        cleaned_chars: list[str] = []
        for ch in text:
            if ch in INVALID_SHEET_CHARS or ch.isspace():
                cleaned_chars.append("_")
            else:
                cleaned_chars.append(ch)
        slug = "".join(cleaned_chars).strip("_")
        return slug or "data"

    def _write_object_sheet(self, ws, obj: dict, context_label: str) -> None:
        ws.cell(row=1, column=1, value="Field").font = self.bold
        ws.cell(row=1, column=2, value="Value").font = self.bold
        if not obj:
            ws.cell(row=2, column=1, value="(no fields)")
            return
        row = 2
        for key, value in obj.items():
            ws.cell(row=row, column=1, value=str(key))
            self._write_value(ws, row, 2, f"{context_label}.{key}", value, field_name=str(key))
            row += 1

    def _write_array_sheet(self, ws, items: list, context_label: str) -> None:
        headers = self._collect_headers(items)
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header).font = self.bold
        for row_idx, item in enumerate(items, start=2):
            if isinstance(item, dict):
                for col_idx, header in enumerate(headers, start=1):
                    self._write_value(
                        ws,
                        row_idx,
                        col_idx,
                        f"{context_label}[{row_idx - 2}].{header}",
                        item.get(header),
                        field_name=str(header),
                    )
            else:
                self._write_value(
                    ws,
                    row_idx,
                    1,
                    f"{context_label}[{row_idx - 2}]",
                    item,
                    field_name="item",
                )

    def _collect_headers(self, items: list) -> list[str]:
        headers: list[str] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            for key in item.keys():
                key_str = str(key)
                if key_str not in headers:
                    headers.append(key_str)
        if not headers:
            headers.append("value")
        return headers

    def _write_value(
        self,
        ws,
        row: int,
        col: int,
        context_label: str,
        value: Any,
        field_name: str,
    ) -> None:
        if isinstance(value, dict):
            if not value:
                ws.cell(row=row, column=col, value="{}")
                return
            sheet = self._create_sheet("OBJ", context_label)
            self._write_object_sheet(sheet, value, context_label)
            self._set_link(ws, row, col, sheet.title, f"{field_name} (object)")
            return
        if isinstance(value, list):
            if not value:
                return  # empty array -> leave parent cell blank
            sheet = self._create_sheet("ARR", context_label)
            self._write_array_sheet(sheet, value, context_label)
            link_text = f"{field_name} ({len(value)} items)"
            self._set_link(ws, row, col, sheet.title, link_text)
            return
        ws.cell(row=row, column=col, value=self._format_scalar(value))

    def _set_link(self, ws, row: int, col: int, sheet_name: str, text: str) -> None:
        cell = ws.cell(row=row, column=col, value=text)
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.style = "Hyperlink"

    def _format_scalar(self, value: Any) -> Any:
        if value is None:
            return "null"
        if isinstance(value, (str, int, float, bool)):
            return value
        return json.dumps(value, ensure_ascii=False)
